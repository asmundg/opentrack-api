"""Competition creation and management for OpenTrack."""

import logging
import tempfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from .browser import OpenTrackSession, screenshot_on_error

logger = logging.getLogger(__name__)

# Normalize young age categories to "10" for OpenTrack import
_CATEGORY_NORMALIZE: dict[str, str] = {
    "Gutter 6-8 Rekrutt": "Gutter 10",
    "Gutter 9": "Gutter 10",
    "Jenter 6-8 Rekrutt": "Jenter 10",
    "Jenter 9": "Jenter 10",
}


def _normalize_xlsx(xlsx_path: Path) -> Path:
    """Normalize young age categories in XLSX and return path to temp copy.

    Rewrites "Klasse" column values like "Gutter 6-8 Rekrutt" and "Gutter 9"
    to "Gutter 10" (and similarly for Jenter) so OpenTrack recognizes them.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Find the "Klasse" column
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    klasse_col = None
    for cell in header_row:
        if cell.value and str(cell.value).strip() == "Klasse":
            klasse_col = cell.column
            break
    if klasse_col is None:
        raise ValueError("No 'Klasse' column found in XLSX")

    # Normalize values
    changed = 0
    for row in ws.iter_rows(min_row=2, min_col=klasse_col, max_col=klasse_col):
        cell = row[0]
        if cell.value and str(cell.value).strip() in _CATEGORY_NORMALIZE:
            cell.value = _CATEGORY_NORMALIZE[str(cell.value).strip()]
            changed += 1

    if changed:
        logger.info("Normalized %d category values in XLSX", changed)

    # Save to a temp file (same suffix so OpenTrack accepts it)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


# Competition type mapping to OpenTrack values
COMPETITION_TYPES = {
    "track": "TRACK",
    "indoor": "INDOOR",
    "road": "ROAD",
    "cross_country": "XC",
    "trail": "TRAIL",
}

# Combined events scoring tables
COMBINED_EVENTS_TABLES = {
    "world_athletics": "WA",
    "tyrving": "TYRV",
}


@dataclass
class CompetitionDetails:
    """Details for creating a new competition."""

    # Required fields
    name: str  # Full name, e.g., "Seriestevne 9-2025"
    slug: str  # URL slug, e.g., "ser9-25"
    start_date: date
    contact_email: str
    organiser_search: str  # Search term to find organiser, e.g., "BULTF"

    # Optional with defaults
    short_name: str = ""  # If empty, OpenTrack may auto-generate
    end_date: date | None = None  # None = same as start_date
    competition_type: Literal["track", "indoor", "road", "cross_country", "trail"] = (
        "track"
    )

    # Display settings
    website: str = ""
    external_entry_link: str = ""  # e.g., Isonen link

    # Scoring settings
    combined_events_table: Literal["world_athletics", "tyrving"] | None = (
        None  # None = use default
    )

    def __post_init__(self):
        if self.end_date is None:
            self.end_date = self.start_date


class CompetitionCreator:
    """Handles creating new competitions on OpenTrack."""

    def __init__(self, session: OpenTrackSession):
        self.session = session
        self.page = session.page

    def create_competition(self, details: CompetitionDetails) -> str:
        """Create a new competition and return its URL.

        Idempotent: if a competition with the given slug already exists at the
        expected URL, skips the create form and resumes configuration. All
        configuration steps are individually idempotent (re-running them on a
        fully configured competition is a no-op).

        Args:
            details: Competition configuration

        Returns:
            URL of the created/resumed competition
        """
        logger.info("Starting competition creation: %s", details.name)

        # Ensure we're logged in
        if not self.session.is_logged_in():
            logger.info("Not logged in, logging in...")
            self.session.login()

        # Step 1: Either create the basic competition or resume an existing one
        if self._find_existing_competition(details):
            logger.info(
                "Step 1/5: Competition '%s' already exists, resuming configuration",
                details.slug,
            )
            self._navigate_to_manage_page()
            self._enable_entries()
        else:
            logger.info("Step 1/5: Creating basic competition...")
            self._create_basic_competition(details)

        # Step 2: Configure advanced settings (hide from public)
        logger.info("Step 2/5: Configuring advanced settings...")
        self._configure_advanced_settings()

        # Step 3: Configure display settings (website, entry link, points, hide results/competitors)
        logger.info("Step 3/5: Configuring display settings...")
        self._configure_display_settings(details)

        # Step 4: Configure scoring (e.g., Tyrving tables)
        if details.combined_events_table:
            logger.info("Step 4/5: Configuring scoring...")
            self._configure_scoring(details)
        else:
            logger.info("Step 4/5: Skipping scoring (not configured)")

        # Step 5: Configure photofinish (FinishLynx)
        logger.info("Step 5/5: Configuring photofinish...")
        self._configure_photofinish()

        logger.info("Competition created successfully: %s", self.page.url)
        return self.page.url

    def _expected_public_url(self, details: CompetitionDetails) -> str:
        """Construct the expected public competition URL from slug + year.

        Norway-specific: assumes the /x/<year>/NOR/<slug>/ path layout.
        """
        base = self.session.config.base_url.rstrip("/")
        year = details.start_date.year
        return f"{base}/x/{year}/NOR/{details.slug}/"

    def _find_existing_competition(self, details: CompetitionDetails) -> bool:
        """Check if a competition with this slug already exists.

        Navigates to the expected public URL and looks for the admin
        " Manage" link as proof we're on a real comp page (and logged in).
        Returns True if found (page is left on the public URL).
        """
        page = self.page
        public_url = self._expected_public_url(details)

        logger.debug("Checking for existing competition at: %s", public_url)
        try:
            response = page.goto(
                public_url, wait_until="domcontentloaded", timeout=30_000
            )
        except Exception as e:
            logger.debug("Navigation to %s failed: %s", public_url, e)
            return False

        if response is None or response.status >= 400:
            status = response.status if response else "?"
            logger.debug("HTTP %s at %s, treating as new competition", status, public_url)
            return False

        # Verify we landed on the expected URL (no redirect to a different comp / home)
        if not page.url.rstrip("/").startswith(public_url.rstrip("/")):
            logger.debug(
                "URL after navigation (%s) does not match expected (%s), treating as new",
                page.url,
                public_url,
            )
            return False

        manage_link = page.get_by_role("link", name=" Manage")
        try:
            manage_link.wait_for(state="visible", timeout=5_000)
        except Exception:
            logger.debug("No Manage link at %s, treating as new competition", public_url)
            return False

        logger.info("Found existing competition at %s", public_url)
        return True

    def _navigate_to_manage_page(self) -> None:
        """From the public competition page, navigate to the Manage page.

        The Manage page is where the settings form (Hide from public, Display,
        Scoring tabs, etc.) and the 'Manage entries' link live. Verifies we
        landed on a page with the expected settings form.
        """
        page = self.page
        page.get_by_role("link", name=" Manage").click()
        page.wait_for_load_state("networkidle")
        # Assert we're on the settings form (has the advanced submit button)
        page.locator('button[name="adv_submit"]').wait_for(
            state="visible", timeout=30_000
        )

    @screenshot_on_error
    def _enable_entries(self) -> None:
        """Enable entries for the competition. Idempotent.

        Captures the current manage page URL up-front. Clicks 'Manage entries';
        if the 'I confirm' dialog appears, accepts it (entries get enabled).
        If not (entries already enabled), navigates back to the captured URL
        so subsequent configuration runs on the right page.
        """
        page = self.page
        manage_url = page.url

        manage_entries = page.get_by_role("link", name="Manage entries")
        try:
            manage_entries.wait_for(state="visible", timeout=5_000)
        except Exception:
            logger.info(
                "'Manage entries' link not present on %s, assuming entries already enabled",
                manage_url,
            )
            return

        logger.debug("Clicking 'Manage entries'")
        manage_entries.click()
        page.wait_for_load_state("networkidle")

        confirm = page.get_by_text("I confirm that I have read")
        try:
            confirm.wait_for(state="visible", timeout=5_000)
        except Exception:
            logger.info(
                "Entries already enabled (no confirmation dialog), returning to %s",
                manage_url,
            )
            page.goto(manage_url)
            page.wait_for_load_state("networkidle")
            page.locator('button[name="adv_submit"]').wait_for(
                state="visible", timeout=30_000
            )
            return

        logger.debug("Confirming entry enablement")
        confirm.click()
        page.get_by_role("button", name="Go").click()
        page.wait_for_load_state("networkidle")
        logger.info("Entries enabled")

    @screenshot_on_error
    def _create_basic_competition(self, details: CompetitionDetails) -> None:
        """Navigate to competition creation and fill basic info."""
        page = self.page

        # Navigate to competition creation
        logger.debug("Navigating to competition creation form")
        page.get_by_role("link", name="Competitions").click()
        page.get_by_role("button", name="New Competition ").click()

        # Fill basic info
        logger.debug("Filling basic competition info")
        page.get_by_role("textbox", name="* Full name:").click()
        page.get_by_role("textbox", name="* Full name:").fill(details.name)

        if details.short_name:
            page.get_by_role("textbox", name="Short name:").fill(details.short_name)

        # Dates
        logger.debug("Setting dates: %s to %s", details.start_date, details.end_date)
        page.get_by_role("textbox", name="* Date:").fill(details.start_date.isoformat())
        if details.end_date and details.end_date != details.start_date:
            page.get_by_role("textbox", name="Finish Date:").fill(
                details.end_date.isoformat()
            )
        else:
            page.get_by_role("textbox", name="Finish Date:").fill(
                details.start_date.isoformat()
            )

        # Slug
        logger.debug("Setting slug: %s", details.slug)
        page.get_by_role("textbox", name="* Slug:").click()
        page.get_by_role("textbox", name="* Slug:").fill(details.slug)

        # Competition type
        comp_type_value = COMPETITION_TYPES.get(details.competition_type, "TRACK")
        logger.debug("Setting competition type: %s", comp_type_value)
        page.get_by_label("Type:").select_option(comp_type_value)

        # Contact email
        logger.debug("Setting contact email: %s", details.contact_email)
        page.get_by_role("textbox", name="* Contact email:").click()
        page.get_by_role("textbox", name="* Contact email:").fill(details.contact_email)

        # Organiser (uses Select2 search widget)
        logger.debug("Selecting organiser: %s", details.organiser_search)
        page.locator("#select2-id_organiser-container").click()
        searchbox = page.get_by_role("searchbox")
        searchbox.fill(details.organiser_search)
        # Wait for the highlighted option to appear and click it
        highlighted_option = page.locator("li.select2-results__option--highlighted")
        highlighted_option.wait_for(state="visible")
        highlighted_option.click()

        # Create the competition. OpenTrack's submit handler may do an AJAX
        # round-trip rather than a classic form POST/redirect, so don't let
        # Playwright block on navigation detection — wait for the post-create
        # page element (the "Manage entries" link) explicitly instead.
        logger.debug("Submitting competition creation form")
        page.get_by_role("button", name="Create").nth(1).click(no_wait_after=True)

        # Wait for the post-create manage page
        logger.debug("Waiting for post-create page")
        page.get_by_role("link", name="Manage entries").wait_for(
            state="visible", timeout=120_000
        )

        # Enable entries (delegates to the idempotent helper)
        self._enable_entries()
        logger.debug("Basic competition created and entries enabled")

    def _configure_advanced_settings(self) -> None:
        """Configure advanced settings - hide competition from public."""
        page = self.page

        logger.debug("Setting competition to hidden from public")
        hide_checkbox = page.get_by_role("checkbox", name="Hide from public:")
        if not hide_checkbox.is_checked():
            hide_checkbox.check()

        logger.debug("Saving advanced settings")
        page.locator('button[name="adv_submit"]').click()
        page.wait_for_load_state("networkidle")

    def _configure_display_settings(self, details: CompetitionDetails) -> None:
        """Configure display settings like website and entry links."""
        page = self.page

        # Navigate to Display tab
        logger.debug("Navigating to Display tab")
        page.get_by_role("link", name="Display ").click()

        # Website
        if details.website:
            logger.debug("Setting website: %s", details.website)
            page.get_by_role("textbox", name="Website:").click()
            page.get_by_role("textbox", name="Website:").fill(details.website)

        # External entry link (e.g., Isonen)
        if details.external_entry_link:
            logger.debug("Setting external entry link: %s", details.external_entry_link)
            page.get_by_role("textbox", name="External entry link:").fill(
                details.external_entry_link
            )

        # Hide results and competitors (useful during setup)
        logger.debug("Hiding results and competitors")
        hide_results = page.get_by_role("checkbox", name="Hide results:")
        if not hide_results.is_checked():
            hide_results.check()
        hide_competitors = page.get_by_role("checkbox", name="Hide competitors:")
        if not hide_competitors.is_checked():
            hide_competitors.check()

        # Save display settings
        logger.debug("Saving display settings")
        page.get_by_role("button", name="Save").nth(1).click()
        page.wait_for_load_state("networkidle")

        # Individual points setting (separate save button)
        logger.debug("Enabling individual points display")
        points_checkbox = page.get_by_role("checkbox", name="Show individual points?")
        if not points_checkbox.is_checked():
            points_checkbox.check()
        page.get_by_role("button", name="Save").nth(3).click()
        page.wait_for_load_state("networkidle")

    def _configure_scoring(self, details: CompetitionDetails) -> None:
        """Configure scoring settings like combined events tables."""
        page = self.page

        # Navigate to Events and scoring. The top-nav link text reflects the
        # currently-configured scoring system: "World Athletics" by default,
        # "Tyrving" once set. Try both to stay idempotent on re-runs.
        logger.debug("Navigating to Events and scoring")
        scoring_nav = None
        for label in ("World Athletics", "Tyrving"):
            candidate = page.locator("a").filter(has_text=label)
            if candidate.count() > 0:
                scoring_nav = candidate.first
                break
        if scoring_nav is None:
            raise RuntimeError("Could not find scoring system navigation link")
        scoring_nav.click()
        page.get_by_role("link", name="Events and scoring").click()

        # Go to Scoring -> Combined Events tab
        logger.debug("Navigating to Combined Events tab")
        page.get_by_role("tab", name="Scoring").click()
        page.get_by_role("tab", name="Combined Events").click()

        # Select the combined events table
        if details.combined_events_table:
            table_value = COMBINED_EVENTS_TABLES.get(
                details.combined_events_table, "WA"
            )
            logger.debug("Setting combined events table: %s", table_value)
            page.get_by_label("Combined Events tables:").select_option(table_value)

        logger.debug("Saving scoring settings")
        page.get_by_role("button", name="Save").click()
        page.wait_for_load_state("networkidle")

    def import_athletes(self, xlsx_path: Path) -> None:
        """Upload XLSX file and process it to create competitor/event records.

        Normalizes young age categories (6-8 Rekrutt, 9 → 10) in a temp copy
        before uploading. Uses OpenTrack's custom import page (manage/custom/)
        with a 3-step workflow: upload → process → fetch PBs (we skip step 3).
        """
        page = self.page

        # Normalize categories in a temp copy
        upload_path = _normalize_xlsx(xlsx_path)
        logger.info("Normalized XLSX written to: %s", upload_path)

        # Navigate to the custom import page
        competition_url = page.url.rstrip("/")
        custom_url = f"{competition_url}/manage/custom/"
        logger.info("Navigating to custom import: %s", custom_url)
        page.goto(custom_url)
        page.wait_for_load_state("networkidle")

        # Step 1: Upload the XLSX file
        logger.info("Uploading file: %s", upload_path.name)
        page.locator("input[name=fileinput]").set_input_files(str(upload_path))
        page.locator("button[name=upload]").click()
        page.wait_for_load_state("networkidle")
        self._wait_for_background_task("upload")

        # Step 2: Process - create competitor/event records
        # Accept the "Are you sure?" confirmation dialog
        logger.info("Processing athletes...")
        page.once("dialog", lambda dialog: dialog.accept())
        page.locator("button[name=process]").click(timeout=240_000)
        page.wait_for_load_state("networkidle")
        self._wait_for_background_task("process")
        logger.info("Athletes imported successfully")

    def prepare_athletes(self) -> None:
        """Number competitors after import."""
        logger.info("Numbering competitors...")
        self._number_competitors()
        logger.info("Athletes prepared successfully")

    def _configure_photofinish(self) -> None:
        """Configure FinishLynx photofinish integration."""
        page = self.page

        # Navigate to TV and photofinish -> FinishLynx
        logger.debug("Navigating to FinishLynx settings")
        page.get_by_role("link", name="TV and photofinish").click()
        page.get_by_role("link", name="FinishLynx").click()

        # Enable photofinish file generation
        logger.debug("Enabling photofinish settings")
        photofinish_files = page.get_by_role(
            "checkbox", name="Photofinish files should"
        )
        if not photofinish_files.is_checked():
            photofinish_files.check()

        # Enable results from photofinish
        results_from_photofinish = page.get_by_role(
            "checkbox", name="Results from photofinish"
        )
        if not results_from_photofinish.is_checked():
            results_from_photofinish.check()

        logger.debug("Saving photofinish settings")
        page.get_by_role("button", name="Save").click()
        page.wait_for_load_state("networkidle")

    def _wait_for_background_task(self, step_name: str) -> None:
        """Wait for a background task to complete if one is running.

        OpenTrack uses a Vue/Pusher component that shows a banner during
        background tasks. The banner container is bg-warning while running
        and changes to bg-success when finished.

        We first wait for bg-warning to appear (task started), then wait
        for it to disappear (task finished, replaced by bg-success).
        """
        page = self.page
        banner = page.locator(".container-fluid.bg-warning")
        # Wait for the task to start (banner appears within a few seconds)
        try:
            banner.wait_for(state="visible", timeout=10_000)
        except Exception:
            return  # No task started
        logger.info("Waiting for %s background task...", step_name)
        banner.wait_for(state="hidden", timeout=120_000)
        logger.info("%s background task complete", step_name)

    def _number_competitors(self) -> None:
        """Automatically assign bib numbers to all competitors."""
        page = self.page

        # Navigate to competitor numbering
        logger.debug("Navigating to competitor numbering")
        page.get_by_role("button", name="Manage competitors ").click()
        page.get_by_role("link", name="Numbering").click()
        page.wait_for_load_state("networkidle")

        # Apply numbering — triggers an AJAX POST, not a full navigation
        logger.debug("Applying bib numbers")
        page.locator("button[name=apply_numbers]").click(no_wait_after=True)
        page.wait_for_load_state("networkidle")
        logger.debug("Bib numbers assigned")

    def _apply_random_seeding(self) -> None:
        """Apply random seeding to all start lists."""
        page = self.page

        # Navigate to seeding
        logger.debug("Navigating to seeding")
        page.get_by_role("button", name="Manage competitors ").click()
        page.get_by_role("link", name="Seeding").click()

        # Fetch all start lists first
        logger.debug("Fetching all start lists")
        page.get_by_text("Fetch all start lists Fetch").click()
        page.wait_for_load_state("networkidle")

        # Apply random order
        logger.debug("Applying random order")
        page.get_by_role("button", name="Random order").click()
        page.wait_for_load_state("networkidle")
        logger.debug("Random seeding applied")
