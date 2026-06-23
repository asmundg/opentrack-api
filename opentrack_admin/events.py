"""Event scheduling and management for OpenTrack competitions."""

import csv
import json
import logging
import re
from dataclasses import dataclass
from datetime import time
from io import StringIO
from pathlib import Path
from typing import Literal
from urllib.parse import urljoin

from pblookup.events import standardize_event_name as pblookup_standardize_event

# Import pblookup for PB/SB lookups
from pblookup.lookup import PBLookupService
from playwright.sync_api import Page

from .browser import OpenTrackSession, screenshot_on_error

# Set up logging
logger = logging.getLogger(__name__)

# Default checkpoint directory
CHECKPOINT_DIR = Path("checkpoints")

# Event code to Norwegian name mapping
EVENT_NAMES = {
    # Jumps
    "HJ": "Høyde",
    "SHJ": "Høyde uten tilløp",
    "LJ": "Lengde",
    "SLJ": "Lengde uten tilløp",
    "TJ": "Tresteg",
    "PV": "Stav",
    # Throws
    "SP": "Kule",
    "DT": "Diskos",
    "JT": "Spyd",
    "HT": "Slegge",
    "BT": "Liten ball",
    # Sprints
    "60m": "60 meter",
    "100m": "100 meter",
    "200m": "200 meter",
    "400m": "400 meter",
    # Middle/Long distance
    "600m": "600 meter",
    "800m": "800 meter",
    "1500m": "1500 meter",
    "3000m": "3000 meter",
    "5000m": "5000 meter",
    # Hurdles
    "60H": "60 meter hekk",
    "80H": "80 meter hekk",
    "100H": "100 meter hekk",
    "110H": "110 meter hekk",
    "200H": "200 meter hekk",
    "400H": "400 meter hekk",
    # Relays
    "4x100m": "4x100 meter",
    "4x400m": "4x400 meter",
}

# Throwing implement weights by category and event
# Format: {event_code: {gender: {age: weight_string}}}
# Gender: "G" for boys/men, "J" for girls/women
# Age groups: 10 (rekrutt/6-10), 11, 12, 13, 14, 15, 16, 17, 20 (U20), 23 (U23), 99 (Senior)
# Note: DT, JT, HT start at age 11 (not offered for rekrutt/10)
# Masters weights are in MASTERS_IMPLEMENT_WEIGHTS — different bracketing.
IMPLEMENT_WEIGHTS = {
    "SP": {  # Kule (Shot Put)
        "G": {
            10: "2",
            11: "2",
            12: "3",
            13: "3",
            14: "4",
            15: "4",
            16: "5",
            17: "5",
            18: "6",
            19: "6",
            20: "7,26",
            23: "7,26",
            99: "7,26",
        },
        "J": {
            10: "2",
            11: "2",
            12: "2",
            13: "2",
            14: "3",
            15: "3",
            16: "3",
            17: "3",
            20: "4",
            23: "4",
            99: "4",
        },
    },
    "DT": {  # Diskos (Discus) - starts at 11
        "G": {
            11: "0,6",
            12: "0,75",
            13: "0,75",
            14: "1",
            15: "1",
            16: "1,5",
            17: "1,5",
            18: "1,75",
            19: "1,75",
            20: "2",
            23: "2",
            99: "2",
        },
        "J": {
            11: "0,6",
            12: "0,6",
            13: "0,6",
            14: "0,75",
            15: "0,75",
            16: "1",
            17: "1",
            20: "1",
            23: "1",
            99: "1",
        },
    },
    "HT": {  # Slegge (Hammer) - starts at 11
        "G": {
            11: "2",
            12: "2",
            13: "3",
            14: "4",
            15: "4",
            16: "5",
            17: "5",
            18: "6",
            19: "6",
            20: "7,26",
            23: "7,26",
            99: "7,26",
        },
        "J": {
            11: "2",
            12: "2",
            13: "2",
            14: "3",
            15: "3",
            16: "3",
            17: "3",
            20: "4",
            23: "4",
            99: "4",
        },
    },
    "JT": {  # Spyd (Javelin) - starts at 11
        # OpenTrack expects kg with Norwegian comma-decimal, same as
        # SP/DT/HT. "400g" was rejected by the form.
        "G": {
            11: "0,4",
            12: "0,4",
            13: "0,4",
            14: "0,6",
            15: "0,6",
            16: "0,7",
            17: "0,7",
            18: "0,8",
            19: "0,8",
            20: "0,8",
            23: "0,8",
            99: "0,8",
        },
        "J": {
            11: "0,4",
            12: "0,4",
            13: "0,4",
            14: "0,4",
            15: "0,4",
            16: "0,5",
            17: "0,5",
            20: "0,5",
            23: "0,6",
            99: "0,6",
        },
    },
}

# Events that use implement weights
THROWING_EVENTS = {"SP", "DT", "HT", "JT"}


# Masters implement weights, keyed by event and gender. Each value is a list
# of (bracket_start_age, weight) sorted ascending; the matching bracket is
# the highest start_age <= athlete_age. Gender "G" = Menn, "J" = Kvinner.
#
# Slegge values omit the wire-length component (e.g. official "7,26/121,5"
# stored as just "7,26") since OpenTrack only takes weight.
# Spyd values are kg with Norwegian comma-decimal (OpenTrack rejects "Ng").
# Vektkast is not modelled — not a standard event in our meet flows.
MASTERS_IMPLEMENT_WEIGHTS: dict[str, dict[str, list[tuple[int, str]]]] = {
    "SP": {
        "G": [(35, "7,26"), (50, "6"), (60, "5"), (70, "4"), (80, "3")],
        "J": [(35, "4"), (50, "3"), (60, "3"), (75, "2")],
    },
    "DT": {
        "G": [(35, "2"), (50, "1,5"), (60, "1"), (70, "1"), (80, "1")],
        "J": [(35, "1"), (50, "1"), (60, "1"), (75, "0,75")],
    },
    "HT": {
        "G": [(35, "7,26"), (50, "6"), (60, "5"), (70, "4"), (80, "3")],
        "J": [(35, "4"), (50, "3"), (60, "3"), (75, "2")],
    },
    "JT": {
        "G": [(35, "0,8"), (50, "0,7"), (60, "0,6"), (70, "0,5"), (80, "0,4")],
        "J": [(35, "0,6"), (50, "0,5"), (60, "0,5"), (75, "0,4")],
    },
}


# Matches the canonical masters short form: "MV45-49", "KV75-79", "MV80",
# etc. Used by both _get_masters_implement_weight and get_implement_weight's
# dispatch (callers should pass normalized categories — normalize_category
# converts long form like "Menn masters 60-64" to "MV60-64").
_MASTERS_CATEGORY_RE = re.compile(r"^([MK])V(\d+)(?:-\d+)?$")


def fold_masters_to_senior(category: str) -> str:
    """Map a masters category (MV*/KV*) to the matching senior (MS/KS).

    OpenTrack imports masters athletes into the senior event pool rather
    than creating separate MV*/KV* events, so any search against OpenTrack
    must look up the corresponding senior event. Non-masters categories
    pass through `normalize_category` unchanged.
    """
    normalized = normalize_category(category)
    m = _MASTERS_CATEGORY_RE.match(normalized)
    if not m:
        return normalized
    return "MS" if m.group(1) == "M" else "KS"


def _get_masters_implement_weight(event_code: str, category: str) -> str | None:
    """Look up a masters implement weight by gender + lower-bound age.

    Returns None if the category doesn't match the masters pattern, the
    event isn't a throw, or the athlete's age is below the lowest defined
    bracket (35).
    """
    m = _MASTERS_CATEGORY_RE.match(normalize_category(category))
    if not m:
        return None

    gender = "G" if m.group(1) == "M" else "J"
    age = int(m.group(2))

    brackets = MASTERS_IMPLEMENT_WEIGHTS.get(event_code, {}).get(gender, [])
    weight: str | None = None
    for bracket_start, w in brackets:
        if age >= bracket_start:
            weight = w
        else:
            break
    return weight


def get_implement_weight(event_code: str, category: str) -> str | None:
    """Get the implement weight for a throwing event and category.

    Args:
        event_code: Event code (SP, DT, HT, JT)
        category: Category like "G10", "J15", "Menn Senior", "Kvinner Senior"

    Returns:
        Weight string (e.g., "2", "0,75", "400g"), or None if event_code is
        not a throwing event, or None if this (event, age) combination is
        intentionally not offered (e.g., DT/JT/HT for rekrutt/10).

    Raises:
        ValueError: If the category is not recognized at all (would otherwise
            silently produce wrong or missing weights).
    """
    if event_code not in THROWING_EVENTS:
        return None

    # Masters categories have their own age-bracketed weight schedule.
    # Detect via normalize_category (handles both raw "Menn masters 60-64"
    # and already-normalized "MV60-64").
    if _MASTERS_CATEGORY_RE.match(normalize_category(category)):
        return _get_masters_implement_weight(event_code, category)

    # Determine gender
    normalized = normalize_category(category)
    if normalized.startswith("G") or normalized in ("M", "MS", "U20", "U23"):
        gender = "G"
    elif normalized.startswith("J") or normalized in ("W", "K", "KS"):
        gender = "J"
    else:
        raise ValueError(
            f"Unknown category {category!r} (normalized={normalized!r}): "
            "cannot determine gender for implement weight lookup"
        )

    # Determine age
    age = get_category_age(category)
    if age is None:
        # Senior/U categories
        if normalized in ("M", "MS", "W", "K", "KS"):
            age = 99
        elif normalized == "U20":
            age = 20
        elif normalized == "U23":
            age = 23
        else:
            raise ValueError(
                f"Unknown category {category!r} (normalized={normalized!r}): "
                "cannot determine age for implement weight lookup"
            )

    # Look up weight
    event_weights = IMPLEMENT_WEIGHTS.get(event_code, {})
    gender_weights = event_weights.get(gender, {})

    # Find the appropriate age bracket
    if age in gender_weights:
        return gender_weights[age]

    # 18-19 athletes fall back to U20 weight when not explicitly listed
    # (e.g., DT/JT only list age 20 for the 18-20 bracket)
    if age in (18, 19) and 20 in gender_weights:
        return gender_weights[20]

    # Age not valid for this event (e.g., DT/JT/HT not offered for rekrutt/10)
    return None


class Checkpoint:
    """Track completed events to allow resuming after failures.

    Stores a simple JSON file with a set of completed event keys.
    """

    def __init__(self, name: str):
        """Create a checkpoint tracker.

        Args:
            name: Name for the checkpoint file (e.g., competition slug)
        """
        CHECKPOINT_DIR.mkdir(exist_ok=True)
        self.path = CHECKPOINT_DIR / f"{name}.json"
        self.completed: set[str] = set()
        self._load()

    def _load(self) -> None:
        """Load existing checkpoint data."""
        if self.path.exists():
            try:
                data = json.loads(self.path.read_text())
                self.completed = set(data.get("completed", []))
                logger.info(
                    f"Loaded checkpoint: {len(self.completed)} events already done"
                )
            except (json.JSONDecodeError, KeyError):
                self.completed = set()

    def _save(self) -> None:
        """Save checkpoint data."""
        self.path.write_text(
            json.dumps({"completed": sorted(self.completed)}, indent=2)
        )

    def is_done(self, key: str) -> bool:
        """Check if an event has been completed."""
        return key in self.completed

    def mark_done(self, key: str) -> None:
        """Mark an event as completed and save."""
        self.completed.add(key)
        self._save()
        logger.debug(f"Checkpoint: marked '{key}' as done")

    def clear(self) -> None:
        """Clear all checkpoint data."""
        self.completed = set()
        if self.path.exists():
            self.path.unlink()


def get_event_name(code: str) -> str:
    """Get the Norwegian event name for an event code.

    Raises:
        KeyError: If the event code is not in the mapping.
    """
    if code in EVENT_NAMES:
        return EVENT_NAMES[code]
    # Handle generic distance codes like "300m" -> "300 meter"
    m = re.match(r"^(\d+)m$", code)
    if m:
        return f"{m.group(1)} meter"
    raise KeyError(
        f"Unknown event code: '{code}'. Add it to EVENT_NAMES in events.py"
    )


# Round/heat annotations OpenTrack appends to entry titles, e.g.
# "G13 HøydePool 1 of 1" or "G/J 11-13 60 meter hekkRace 1 of 1".
_EVENT_TITLE_SUFFIX_RE = re.compile(
    r"\s*(Race|Pool|Heat|Final|Round|Group)\s*\d.*$"
)

# Known discipline names, longest first so "60 meter hekk" beats "60 meter"
# and "Høyde uten tilløp" beats "Høyde" when matching a title suffix.
_EVENT_NAMES_BY_LENGTH = sorted(set(EVENT_NAMES.values()), key=len, reverse=True)
_EVENT_NAME_TO_CODE_EXACT = {v: k for k, v in EVENT_NAMES.items()}


def event_code_from_event_title(title: str) -> str | None:
    """Derive the discipline code from an OpenTrack event title.

    Event titles combine a category prefix with the discipline name, e.g.
    "G/J 11-13 60 meter hekk" or a merged "G 12-14 600 meter". A trailing round
    annotation ("...Pool 1 of 1") is stripped first. The discipline is matched
    as the title's suffix against the known event-name vocabulary (longest match
    wins); generic distances fall back to a numeric pattern. Returns the code
    (e.g. "60H", "600m", "HJ") or None when no discipline can be identified.
    """
    cleaned = _EVENT_TITLE_SUFFIX_RE.sub("", title).strip()
    for name in _EVENT_NAMES_BY_LENGTH:
        if cleaned.endswith(name):
            return _EVENT_NAME_TO_CODE_EXACT[name]
    m = re.search(r"(\d+) meter hekk$", cleaned)
    if m:
        return f"{m.group(1)}H"
    m = re.search(r"(\d+) meter$", cleaned)
    if m:
        return f"{m.group(1)}m"
    return None


# Map any known category alias to its canonical OpenTrack short code.
# Covers Isonen XLSX values ("Gutter 14"), scheduler enum values
# ("Kvinner Senior", "J-Rekrutt"), and lowercase / casing variants. Lookup
# is case-insensitive — see normalize_category(). Already-canonical codes
# ("G14", "KS", ...) pass through via the .get() fallback.
_CATEGORY_ALIASES: dict[str, str] = {
    # Boys — Isonen names
    "gutter 6-8 rekrutt": "G10",
    "gutter 9": "G10",
    "gutter 10": "G10",
    "gutter 11": "G11",
    "gutter 12": "G12",
    "gutter 13": "G13",
    "gutter 14": "G14",
    "gutter 15": "G15",
    "gutter 16": "G16",
    "gutter 17": "G17",
    "gutter 18-19": "G18-19",
    # Girls — Isonen names
    "jenter 6-8 rekrutt": "J10",
    "jenter 9": "J10",
    "jenter 10": "J10",
    "jenter 11": "J11",
    "jenter 12": "J12",
    "jenter 13": "J13",
    "jenter 14": "J14",
    "jenter 15": "J15",
    "jenter 16": "J16",
    "jenter 17": "J17",
    "jenter 18-19": "J18-19",
    # Seniors — both Isonen and scheduler use these long names
    "kvinner senior": "KS",
    "menn senior": "MS",
    # Scheduler enum shortcuts
    "g-rekrutt": "G10",
    "j-rekrutt": "J10",
}


# Matches Isonen long-form masters categories like "Menn masters 60-64",
# normalized to canonical "MV60-64" / "KV60-64" by normalize_category().
_MASTERS_LONGFORM_RE = re.compile(
    r"^(menn|kvinner)\s+masters\s+(\d+(?:-\d+)?)$",
    re.IGNORECASE,
)


def normalize_category(category: str) -> str:
    """Map any known category alias to its canonical OpenTrack short code.

    Recognizes (case-insensitively):
      - Isonen XLSX values: "Gutter 14" -> "G14", "Kvinner Senior" -> "KS"
      - Scheduler enum values: "J-Rekrutt" -> "J10", "Menn Senior" -> "MS"
      - Masters long form: "Menn masters 60-64" -> "MV60-64",
        "Kvinner masters 75-79" -> "KV75-79"
      - Already-canonical codes ("G14", "J18-19", "KS", "MS", "MV60-64") pass through.

    This is the single normalization mechanism used by both the schedule
    parser (Isonen XLSX/CSV → EventSchedule) and the OpenTrack search code
    (EventSchedule → search query). Anything not in _CATEGORY_ALIASES is
    returned unchanged so callers like the OpenTrack search can fail fast
    with a clear "no match" error.
    """
    m = _MASTERS_LONGFORM_RE.match(category.strip())
    if m:
        prefix = "MV" if m.group(1).lower() == "menn" else "KV"
        return f"{prefix}{m.group(2)}"
    return _CATEGORY_ALIASES.get(category.lower(), category)


def get_category_age(category: str) -> int | None:
    """Extract age from category like 'G10', 'J15', 'G18-19', etc.

    Returns None for senior categories like 'M', 'W', 'U20', 'U23'.
    For ranges like 'G18-19', returns the lower bound (18).
    """
    normalized = normalize_category(category)
    # Match G/J followed by age number, optionally with a range like "18-19"
    match = re.match(r"^[GJ](\d+)(?:-\d+)?$", normalized)
    if match:
        return int(match.group(1))
    # Senior/U categories treated as 15+
    return None


def is_field_event(event_code: str) -> bool:
    """Check if event code is a field event (jumps/throws)."""
    return event_code in {"HJ", "SHJ", "LJ", "SLJ", "TJ", "PV", "SP", "DT", "JT", "HT"}


def is_track_event(event_code: str) -> bool:
    """Check if an event code is a track (running) event.

    Track events are sprints/distance ("60m".."5000m"), hurdles ("60H".."400H")
    and relays ("4x100m"). Everything else (jumps, throws, and the "Liten ball"
    ball-throw) is a field event. Defined positively so unknown codes are not
    mistaken for track heats.
    """
    return bool(
        re.match(r"^\d+(m|H)$", event_code) or re.match(r"^\d+x\d+m$", event_code)
    )


def is_horizontal_field_event(event_code: str) -> bool:
    """Check if event code is a horizontal field event (has attempts).

    Horizontal jumps and throws have a fixed number of attempts.
    Vertical jumps (HJ, SHJ, PV) are height-based and don't use attempts.
    """
    return event_code in {"LJ", "SLJ", "TJ", "SP", "DT", "JT", "HT"}


@dataclass
class AttemptConfig:
    """Configuration for number of attempts in field events.

    Based on Norwegian athletics rules:
    - FIFA and rekrutt (6-10): 3 attempts, no cut
    - 11-12 years: 4 attempts, no cut
    - 13+ and seniors: 6 attempts with cut after 3 (top 8 continue)
    """

    attempts: int  # Total number of attempts
    field_cut: int  # Cut after this many attempts (0 = no cut)

    @classmethod
    def for_category(cls, category: str) -> "AttemptConfig":
        """Get attempt configuration for a category."""
        age = get_category_age(category)

        if age is not None and age <= 10:
            # FIFA and rekrutt: 3 attempts, no cut
            return cls(attempts=3, field_cut=0)
        elif age is not None and age <= 12:
            # 11-12 years: 4 attempts, no cut
            return cls(attempts=4, field_cut=0)
        else:
            # 13+ and seniors: 6 attempts, cut after 3
            return cls(attempts=6, field_cut=3)


@dataclass
class EventSchedule:
    """A scheduled event with category, event code, and start time.

    Examples:
        EventSchedule("G10", "60m", time(10, 0))     # G10 60m at 10:00
        EventSchedule("J15", "HJ", time(10, 30))     # J15 Høyde at 10:30
        EventSchedule("M", "SP", time(11, 0))        # Men Kule at 11:00
    """

    category: str  # e.g., "G10", "J15", "M", "W", "G-Rekrutt"
    event: str  # e.g., "60m", "HJ", "SP", "LJ"
    start_time: time

    @property
    def event_name(self) -> str:
        """Get the Norwegian name for the event."""
        return get_event_name(self.event)

    @property
    def search_category(self) -> str:
        """Get the category to use when searching for an OpenTrack event.

        Masters (MV*/KV*) are folded to MS/KS because OpenTrack imports
        masters athletes into the senior event pool. See
        `fold_masters_to_senior`.
        """
        return fold_masters_to_senior(self.category)

    @property
    def search_term(self) -> str:
        """Generate search term for finding this event."""
        return f"{self.search_category} {self.event_name}"

    @property
    def is_field_event(self) -> bool:
        """Check if this is a field event (jumps/throws)."""
        return is_field_event(self.event)

    @property
    def is_horizontal_field_event(self) -> bool:
        """Check if this is a horizontal field event (has attempts).

        HJ and PV are vertical and height-based, no attempts.
        """
        return is_horizontal_field_event(self.event)

    @property
    def attempt_config(self) -> AttemptConfig:
        """Get the attempt configuration for this event's category."""
        return AttemptConfig.for_category(self.category)

    @property
    def is_throwing_event(self) -> bool:
        """Check if this is a throwing event (has implement weights)."""
        return self.event in THROWING_EVENTS

    @property
    def implement_weight(self) -> str | None:
        """Get the implement weight for this throwing event, or None."""
        return get_implement_weight(self.event, self.category)


@dataclass
class EventMergeGroup:
    """Track events of one discipline that share a heat and must be merged.

    A schedule_events.csv row groups the categories that run together. For track
    events that means a single heat, which in OpenTrack is expressed by merging
    the per-category events into one. `primary` is the event kept in OpenTrack;
    every event in `others` is merged into it.
    """

    primary: EventSchedule
    others: list[EventSchedule]

    @property
    def members(self) -> list[EventSchedule]:
        """All events in the group, primary first."""
        return [self.primary, *self.others]

    @property
    def merged_name(self) -> str:
        """Combined name for the merged event, e.g. "G/J 13-14 200 meter".

        Format: ``<gender> [<age range>] <event name>``. Gender is G (boys),
        J (girls) or G/J (both) for youth, or M/K for senior-only groups. The
        age range spans the members' ages (a single number when they coincide);
        senior members contribute no age. All members share one event code, so
        the event name is taken from the primary.
        """
        normalized = [normalize_category(m.category) for m in self.members]
        ages = sorted(
            age
            for age in (get_category_age(m.category) for m in self.members)
            if age is not None
        )
        has_male = any(
            n.startswith("G") or n == "MS" or n.startswith("MV") for n in normalized
        )
        has_female = any(
            n.startswith("J") or n == "KS" or n.startswith("KV") for n in normalized
        )
        male_letter, female_letter = ("G", "J") if ages else ("M", "K")
        if has_male and has_female:
            gender = f"{male_letter}/{female_letter}"
        elif has_female:
            gender = female_letter
        else:
            gender = male_letter

        event_name = get_event_name(self.primary.event)
        if ages:
            lo, hi = ages[0], ages[-1]
            age_range = str(lo) if lo == hi else f"{lo}-{hi}"
            return f"{gender} {age_range} {event_name}"
        return f"{gender} {event_name}"


def lookup_athlete_pb(
    service: PBLookupService,
    name: str,
    club: str,
    birth_date: str,
    event: str,
    category: str,
    debug: bool = False,
) -> str | None:
    """Look up one athlete's formatted PB for a discipline, or ``None``.

    ``event`` is an admin discipline code (e.g. ``"HJ"``, ``"100m"``, ``"60H"``);
    it is translated into the pblookup vocabulary before searching. Returns the
    OpenTrack-ready result string (e.g. ``"10.54"``, ``"6:00.80"``) or ``None``
    when no result is found. Shared by the browser and API update-pbs paths.
    """
    pblookup_event = pblookup_standardize_event(EVENT_NAMES.get(event, event))
    try:
        result = service.lookup_pb(
            name, club, birth_date, pblookup_event, category=category
        )
    except Exception as e:
        logger.warning(f"Error looking up PB for {name}: {e}")
        return None
    if result is None:
        return None
    return result.get_result_formatted()


class EventScheduler:
    """Handles scheduling events in an OpenTrack competition."""

    def __init__(self, session: OpenTrackSession):
        self.session = session
        self.page = session.page

    @screenshot_on_error
    def navigate_to_events_table(self) -> None:
        """Navigate to the Events Table tab from anywhere."""
        page = self.page

        # Navigate through Manage -> Events and scoring -> Events Table
        logger.info("Navigating to Events Table...")
        page.get_by_role("link", name=" Manage").click()
        page.get_by_role("link", name="Events and scoring").click()

        # Wait for the events form to load before clicking the tab
        page.locator("form").filter(has_text="IDNumCodeAge").wait_for(state="visible")
        page.get_by_role("tab", name=" Events Table").click()
        page.wait_for_load_state("networkidle")

        logger.info("Events table ready")

    @screenshot_on_error
    def find_and_click_event(self, schedule: EventSchedule) -> bool:
        """Find an event by category and code, then click to open it.

        Args:
            schedule: The event to find

        Returns:
            True if event was found and clicked, False otherwise
        """
        page = self.page

        if schedule.search_category != normalize_category(schedule.category):
            logger.info(
                "Searching for event: %s (folded from %s)",
                schedule.search_term,
                schedule.category,
            )
        else:
            logger.info("Searching for event: %s", schedule.search_term)

        # Use the search box to filter
        search_box = page.get_by_role("textbox", name="Search")
        search_box.click()
        search_box.fill(schedule.search_term)

        # Click the search button
        page.locator("#search_advanced_btn").click()

        # Wait for table to filter
        page.wait_for_load_state("networkidle")

        # Find matching rows by checking the event name column for an exact match.
        # The search box returns partial matches (e.g., "Høyde" matches both
        # "Høyde" and "Høyde uten tilløp"), so we must verify the name cell.
        expected_name = schedule.search_term  # e.g., "J18-19 Høyde"
        rows = page.locator("tr").filter(
            has=page.locator("a").filter(has_text=re.compile(r"^[TF]\d+$"))
        )

        count = rows.count()
        logger.info(f"Found {count} candidate row(s) for '{expected_name}'")

        for i in range(count):
            row = rows.nth(i)
            name_cell = row.locator("td[data-mdb-field='name']")
            if name_cell.count() == 0:
                continue
            row_name = (name_cell.text_content() or "").strip()
            if row_name == expected_name:
                link = row.locator("a").filter(has_text=re.compile(r"^[TF]\d+$")).first
                logger.info(f"Exact match: '{row_name}' — clicking {link.text_content()}")
                link.click()
                page.wait_for_load_state("networkidle")
                return True

        # No exact match found - raise error (screenshot taken by wrapper)
        found_names = []
        for i in range(count):
            name_cell = rows.nth(i).locator("td[data-mdb-field='name']")
            if name_cell.count() > 0:
                found_names.append((name_cell.text_content() or "").strip())
        raise RuntimeError(
            f"No exact match for '{expected_name}'. Found: {found_names}"
        )

    @screenshot_on_error
    def set_implement_weights(self, event_code: str) -> None:
        """Set per-competitor implement weights for a throwing event.

        Navigates to the per-pool attempts/heights editor, reads each
        athlete row's category from the Handsontable data (the 'category'
        column is hidden in the UI but present in the data model), and
        types the resolved weight from `get_implement_weight(event_code,
        row_category)` into the Weight column.

        Per-row resolution is required because masters athletes (MV*/KV*)
        are folded into senior pools (MS/KS) on OpenTrack, but they use
        their own age-bracketed implement weights — not the senior weight.

        Assumes we're on the event detail page. Pool seeding must have
        been done first (athletes need QPs and to be assigned to pools);
        otherwise the edit table will be empty. Run `opentrack admin
        update-pbs` before scheduling throwing events on a fresh import.

        Args:
            event_code: OpenTrack event code (SP, DT, HT, JT)

        Raises:
            RuntimeError: If the pool is empty, the data table is
                unreadable, or any row's category yields no weight for
                this event (unknown category or category not offered for
                this event).
        """
        page = self.page

        logger.info(f"Setting implement weights for event {event_code}")

        # After set_event_attempts we're on /manage/events/{code}/, which has
        # no per-pool Edit link. Navigate to the event detail view first to
        # find the pool-specific edit URL.
        view_link = page.get_by_role("link", name=re.compile(r"^VIEW$", re.IGNORECASE))
        view_href = view_link.first.get_attribute("href")
        if not view_href:
            raise RuntimeError("Could not find VIEW link on Manage event page")
        page.goto(urljoin(page.url, view_href), wait_until="load")

        # The per-pool Edit link is rendered inside a responsive button group
        # (collapse_right) that is hidden on narrow viewports, so clicking it
        # times out. Read its href and navigate directly instead.
        edit_link = page.locator("a[href*='/edit/']").first
        edit_href = edit_link.get_attribute("href")
        if not edit_href:
            raise RuntimeError("Could not find per-pool Edit link on event page")
        edit_url = urljoin(page.url, edit_href)
        logger.info(f"Navigating to weight editor: {edit_url}")
        page.goto(edit_url, wait_until="load")

        # Wait for Handsontable to render before reading rows.
        page.locator("#attempts_ht .ht_master table.htCore").wait_for(state="visible")

        # Read the underlying data array (includes hidden 'category' column).
        rows_data = page.evaluate(
            """() => {
                const el = document.querySelector('#attempts_ht');
                if (!el) return null;
                let inst = null;
                if (window.$ && $(el).handsontable) {
                    try { inst = $(el).handsontable('getInstance'); } catch (e) {}
                }
                if (!inst && el.hotInstance) inst = el.hotInstance;
                if (!inst && window.Handsontable && Handsontable.getInstance) {
                    try { inst = Handsontable.getInstance(el); } catch (e) {}
                }
                if (!inst) return null;
                return inst.getData();
            }"""
        )
        if not rows_data:
            raise RuntimeError("Could not read Handsontable data from weight editor")

        populated = [r for r in rows_data if (r.get("bib") or "").strip()]
        if not populated:
            raise RuntimeError(
                "Pool is empty (no athletes seeded). Run 'opentrack admin "
                "update-pbs' to populate seeding performances before scheduling "
                "throwing events."
            )

        # Resolve all weights up front so we fail fast on bad categories
        # before mutating any cells.
        per_row_weights: list[str] = []
        for row in populated:
            category = (row.get("category") or "").strip()
            if not category:
                raise RuntimeError(
                    f"Athlete bib {row.get('bib')!r} "
                    f"({row.get('first_name')} {row.get('last_name')}) "
                    "has no category in pool data; cannot resolve implement weight."
                )
            try:
                weight = get_implement_weight(event_code, category)
            except ValueError as e:
                raise RuntimeError(
                    f"Cannot resolve implement weight for athlete bib "
                    f"{row.get('bib')!r} ({row.get('first_name')} "
                    f"{row.get('last_name')}, category {category!r}) "
                    f"in event {event_code}: {e}"
                ) from e
            if weight is None:
                raise RuntimeError(
                    f"No implement weight defined for athlete bib "
                    f"{row.get('bib')!r} ({row.get('first_name')} "
                    f"{row.get('last_name')}, category {category!r}) "
                    f"in event {event_code}."
                )
            per_row_weights.append(weight)

        # Find the Weight column header to locate the column's X position.
        table = page.locator("#attempts_ht .ht_master table.htCore")
        weight_header = table.locator("thead th").filter(has_text="Weight")
        if weight_header.count() == 0:
            raise RuntimeError("Could not find 'Weight' column in table")

        weight_header_box = weight_header.bounding_box()
        if not weight_header_box:
            raise RuntimeError("Could not get Weight column position")
        weight_x = weight_header_box["x"] + weight_header_box["width"] / 2

        rows = table.locator("tbody tr")
        logger.info(
            "Filling weights for %d athlete row(s): %s",
            len(populated),
            ", ".join(
                f"{r.get('bib')}={w}" for r, w in zip(populated, per_row_weights)
            ),
        )

        for i, weight in enumerate(per_row_weights):
            row_box = rows.nth(i).bounding_box()
            if not row_box:
                continue
            row_y = row_box["y"] + row_box["height"] / 2

            page.mouse.click(weight_x, row_y)
            page.wait_for_timeout(100)
            page.keyboard.type(weight)
            page.keyboard.press("Enter")
            logger.debug(f"Set weight for row {i + 1}: {weight}")

        # The Save button POSTs the form and reloads the edit page (no
        # in-page banner like the event-detail form). Wait for navigation.
        with page.expect_navigation(wait_until="load"):
            page.get_by_role("button", name="Save").first.click()
        logger.info("Implement weights saved")

    @screenshot_on_error
    def merge_event_group(self, group: EventMergeGroup) -> None:
        """Merge a group's sibling track events into its primary event.

        Opens the primary event's detail page and drives OpenTrack's merge form
        (``#eventMergeForm``): it ticks the checkbox for each sibling (matched by
        the Name column, which equals the event's ``search_term``) and submits,
        accepting the irreversible-merge confirmation dialog.

        OpenTrack performs the merge as a queued background task, so after
        submitting we poll the candidate list until the merged siblings are gone.

        Idempotent: siblings that are no longer offered as candidates are assumed
        already merged (e.g. on a resumed run) and skipped; if nothing remains to
        merge the call is a no-op. Assumes the primary still exists in OpenTrack.
        """
        page = self.page

        logger.info(
            "=== Merging into %s: %s ===",
            group.primary.search_term,
            [s.search_term for s in group.others],
        )
        self.find_and_click_event(group.primary)
        detail_url = page.url

        if page.locator("#eventMergeForm").count() == 0:
            # OpenTrack only renders the merge form when same-discipline
            # candidates exist; its absence means there is nothing left to merge
            # into the primary (e.g. the group was already merged on a prior run).
            logger.info(
                "No merge form on %s; nothing left to merge",
                group.primary.search_term,
            )
            return

        candidates = self._merge_candidate_terms()
        to_merge = [s for s in group.others if s.search_term in candidates]
        for sib in group.others:
            if sib.search_term not in candidates:
                logger.info(
                    "'%s' is not a merge candidate for '%s' (already merged?), "
                    "skipping",
                    sib.search_term,
                    group.primary.search_term,
                )

        if not to_merge:
            logger.info(
                "Nothing to merge into %s; all siblings already merged",
                group.primary.search_term,
            )
            return

        for sib in to_merge:
            checkbox = page.locator("#eventMergeForm tbody tr").filter(
                has=page.get_by_role("cell", name=sib.search_term, exact=True)
            ).locator("input[type='checkbox']")
            if checkbox.count() != 1:
                raise RuntimeError(
                    f"Merge candidate '{sib.search_term}' is not unique on "
                    f"'{group.primary.search_term}' page; found {checkbox.count()}."
                )
            checkbox.check()

        # Apply the combined name (e.g. "G/J 13-14 200 meter") so the merged
        # event reflects all its categories instead of keeping the primary's.
        merged_name = group.merged_name
        logger.info("Naming merged event: %s", merged_name)
        page.locator("#eventMergeForm #id_name").fill(merged_name)

        # The visible Submit button (#event-merge) is type=button; its JS shows a
        # confirm() then clicks the hidden submit. Accept the dialog, then wait
        # for the POST to reload the detail page.
        page.once("dialog", lambda d: d.accept())
        page.locator("#event-merge").click()
        page.wait_for_load_state("networkidle")

        # The merge is processed asynchronously as a queued background task; wait
        # for it to finish (siblings drop off the candidate list).
        self._wait_for_merge(detail_url, [s.search_term for s in to_merge])

        logger.info(
            "Merged %d event(s) into %s",
            len(to_merge),
            group.primary.search_term,
        )

    def _merge_candidate_terms(self) -> set[str]:
        """Return the Name cells currently offered in the merge form."""
        names = self.page.evaluate(
            "() => Array.from(document.querySelectorAll("
            "'#eventMergeForm tbody tr td:nth-child(6)')).map(td => td.textContent.trim())"
        )
        return set(names)

    def _wait_for_merge(
        self,
        detail_url: str,
        merged_terms: list[str],
        attempts: int = 40,
        interval_ms: int = 3000,
    ) -> None:
        """Poll the primary's merge form until ``merged_terms`` are gone.

        OpenTrack queues the merge as a background task, so the siblings stay
        listed as candidates for a few seconds after submit. Reloads the detail
        page between checks. Raises if they are still present after the timeout.
        """
        page = self.page
        remaining = [t for t in merged_terms if t in self._merge_candidate_terms()]
        for attempt in range(1, attempts + 1):
            if not remaining:
                return
            logger.debug(
                "Waiting for merge background task (still listed: %s, %d/%d)",
                remaining,
                attempt,
                attempts,
            )
            page.wait_for_timeout(interval_ms)
            page.goto(detail_url, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")
            remaining = [t for t in merged_terms if t in self._merge_candidate_terms()]
        if remaining:
            raise RuntimeError(
                f"Merge at {detail_url} did not complete; still listed: {remaining}"
            )

    def merge_event_groups(
        self,
        groups: list[EventMergeGroup],
        checkpoint_name: str | None = None,
    ) -> None:
        """Merge every track merge-group, with optional checkpoint resume.

        Must run AFTER ``schedule_events`` so that all member events exist with
        their start times. Checkpoint keys are namespaced (``merge:<term>``) so
        the same file can be shared with the scheduling pass.
        """
        if not groups:
            return

        checkpoint = Checkpoint(checkpoint_name) if checkpoint_name else None
        logger.info("Merging %d track event group(s)", len(groups))

        for i, group in enumerate(groups, 1):
            key = f"merge:{group.primary.search_term}"
            if checkpoint and checkpoint.is_done(key):
                logger.info(
                    "Skipping merge %d/%d: %s (already done)", i, len(groups), key
                )
                continue

            logger.info("Processing merge group %d/%d", i, len(groups))
            self.navigate_to_events_table()
            self.merge_event_group(group)

            if checkpoint:
                checkpoint.mark_done(key)

        logger.info("Finished merging %d group(s)", len(groups))


def parse_time(time_str: str) -> time:
    """Parse a time string like '17:00' or '17:25'."""
    hour, minute = map(int, time_str.strip().split(":"))
    return time(hour, minute)


def parse_schedule_file(path: Path) -> list[EventSchedule]:
    """Parse a schedule file, auto-detecting the format.

    Supports:
    - .xlsx → Isonen-format spreadsheet (via parse_schedule_xlsx)
    - .csv with "event_type" header → scheduler event-overview CSV
      (via parse_event_schedule_csv)
    - other .csv → Isonen-format participant CSV (via parse_schedule_csv)
    """
    if path.suffix.lower() == ".xlsx":
        return parse_schedule_xlsx(path)

    # CSV: peek at the first line to distinguish event-overview from Isonen.
    header = path.read_text().split("\n", 1)[0]
    if "event_type" in header:
        return parse_event_schedule_csv(path)
    return parse_schedule_csv(path.read_text())


def parse_schedule_csv(content: str) -> list[EventSchedule]:
    """Parse an Isonen-format schedule CSV file.

    Expected format (schedule.csv from scheduler):
        Fornavn,Etternavn,...,Klasse,Øvelse,...,Kl.,...
        John,Doe,...,Gutter 14,Lengde,...,17:00,...

    Args:
        content: CSV file content as string

    Returns:
        List of EventSchedule objects (deduplicated by category+event)
    """
    schedules = []
    seen: set[tuple[str, str]] = set()  # For deduplication

    reader = csv.DictReader(StringIO(content))

    for row in reader:
        try:
            category = normalize_category(row["Klasse"].strip())
            event = _normalize_isonen_event(row["Øvelse"].strip())
            start_time = parse_time(row["Kl."])

            # Skip if we've already seen this category+event combination
            key = (category, event)
            if key in seen:
                continue
            seen.add(key)

            schedules.append(
                EventSchedule(
                    category=category,
                    event=event,
                    start_time=start_time,
                )
            )
        except (KeyError, ValueError):
            # Skip invalid rows
            continue

    return schedules


def parse_schedule_xlsx(xlsx_path: Path) -> list[EventSchedule]:
    """Parse an Isonen-format XLSX file to extract unique event schedules.

    Same logic as parse_schedule_csv but reads from XLSX directly.
    Expected columns: Klasse, Øvelse (Kl. is optional, defaults to 00:00).
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows()

    header_row = next(rows_iter)
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]

    schedules: list[EventSchedule] = []
    seen: set[tuple[str, str]] = set()

    for row in rows_iter:
        values = {h: (str(c.value).strip() if c.value is not None else "") for h, c in zip(headers, row)}
        try:
            category = normalize_category(values["Klasse"])
            event = _normalize_isonen_event(values["Øvelse"])
            start_time = parse_time(values["Kl."]) if values.get("Kl.") else time(0, 0)

            key = (category, event)
            if key in seen:
                continue
            seen.add(key)

            schedules.append(EventSchedule(category=category, event=event, start_time=start_time))
        except (KeyError, ValueError):
            continue

    wb.close()
    return schedules


# Mapping from Isonen event names to event codes
ISONEN_EVENT_CODES: dict[str, str] = {
    # Track events
    "60 meter": "60m",
    "100 meter": "100m",
    "200 meter": "200m",
    "400 meter": "400m",
    "600 meter": "600m",
    "800 meter": "800m",
    "1500 meter": "1500m",
    "3000 meter": "3000m",
    "5000 meter": "5000m",
    # Hurdles
    "60 meter hekk": "60H",
    "80 meter hekk": "80H",
    "100 meter hekk": "100H",
    "110 meter hekk": "110H",
    "400 meter hekk": "400H",
    # Field events
    "Høyde": "HJ",
    "Høyde uten tilløp": "SHJ",
    "Lengde": "LJ",
    "Lengde uten tilløp": "SLJ",
    "Tresteg": "TJ",
    "Stav": "PV",
    "Stavsprang": "PV",
    "Kule": "SP",
    "Diskos": "DT",
    "Spyd": "JT",
    "Slegge": "HT",
    "Liten ball": "BT",  # Ball throw for young athletes
}


def _normalize_isonen_event(event: str) -> str:
    """Convert Isonen event name to event code.

    E.g., "60 meter" -> "60m", "Lengde" -> "LJ"
    """
    if event in ISONEN_EVENT_CODES:
        return ISONEN_EVENT_CODES[event]
    # Handle "NNN meter" / "NNN meter hekk" patterns generically
    m = re.match(r"^(\d+) meter hekk$", event)
    if m:
        return f"{m.group(1)}H"
    m = re.match(r"^(\d+) meter$", event)
    if m:
        return f"{m.group(1)}m"
    return event


# Reverse mapping from Norwegian event names (as used in scheduler EventType values)
# to admin event codes. Covers both Isonen names and scheduler-specific variants.
_EVENT_NAME_TO_CODE: dict[str, str] = {v: k for k, v in EVENT_NAMES.items()}
_EVENT_NAME_TO_CODE.update({
    # Scheduler uses different names than EVENT_NAMES for some events
    "Stavsprang": "PV",
    "Liten ball": "BT",
    "60m hekk": "60H",
    "80m hekk": "80H",
    "100m hekk": "100H",
    "200m hekk": "200H",
})


def parse_event_schedule_csv(path: Path) -> list[EventSchedule]:
    """Parse a schedule_events.csv (event overview from scheduler).

    Each row has combined categories (e.g., "G14,J14") which are split
    into individual EventSchedule entries sharing the same start time.

    Duplicate (category, event) pairs are silently dropped (first occurrence wins).

    Expected columns: event_type, categories, start_time
    """
    schedules: list[EventSchedule] = []
    seen: set[tuple[str, str]] = set()

    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for line_num, row in enumerate(reader, start=2):
            event_name = row["event_type"].strip()
            # Look up event code, fall back to identity (handles "60m" etc.)
            event_code = _EVENT_NAME_TO_CODE.get(event_name, event_name)
            start = parse_time(row["start_time"])

            for cat in row["categories"].split(","):
                cat = cat.strip()
                key = (cat, event_code)
                if key in seen:
                    continue
                seen.add(key)
                schedules.append(EventSchedule(
                    category=cat,
                    event=event_code,
                    start_time=start,
                ))

    return schedules


def parse_event_merge_groups(path: Path) -> list[EventMergeGroup]:
    """Parse a schedule_events.csv into track-event merge groups.

    Each CSV row groups the categories of one event type that run together. For
    a TRACK event, several categories in a row means a shared heat, expressed in
    OpenTrack by merging the per-category events into one. Non-track rows (jumps,
    throws, ball-throw) are not merged here (they share a venue but run as
    separate competitions).

    Returns one EventMergeGroup per track row that resolves to at least two
    distinct OpenTrack events (after masters-folding via ``search_term``). The
    first category in the row is the primary; the rest are merged into it.

    Expected columns: event_type, categories, start_time.
    """
    groups: list[EventMergeGroup] = []

    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            event_name = row["event_type"].strip()
            event_code = _EVENT_NAME_TO_CODE.get(event_name, event_name)
            if not is_track_event(event_code):
                continue

            start = parse_time(row["start_time"])
            members: list[EventSchedule] = []
            seen_terms: set[str] = set()
            for cat in row["categories"].split(","):
                cat = cat.strip()
                if not cat or cat.upper() == "FIFA":
                    continue
                sched = EventSchedule(category=cat, event=event_code, start_time=start)
                # Distinct OpenTrack events only: categories that fold to the
                # same search term (e.g. masters folded to senior) are one event.
                if sched.search_term in seen_terms:
                    continue
                seen_terms.add(sched.search_term)
                members.append(sched)

            if len(members) >= 2:
                groups.append(
                    EventMergeGroup(primary=members[0], others=members[1:])
                )

    return groups
