"""Export Isonen CSV to OpenTrack XLSX format for event creation."""

import csv
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class EventRow(NamedTuple):
    """Represents a unique event to be created in OpenTrack."""
    category: str  # e.g., "G10", "J15"
    event: str     # e.g., "60m", "LJ", "SP"


def collapse_category(category: str) -> str:
    """Collapse categories <=10 into '10'.

    Examples:
        "Gutter 9" -> "Gutter 10"
        "Gutter Rekrutt 6-8" -> "Gutter 10"
        "Jenter 9" -> "Jenter 10"
        "Jenter Rekrutt 6-8" -> "Jenter 10"
        "Gutter 15" -> "Gutter 15" (unchanged)

    Args:
        category: Original category from Isonen CSV

    Returns:
        Collapsed category with <=10 mapped to 10
    """
    # Normalize to lowercase for comparison
    cat_lower = category.lower()

    # Check if it's a rekrutt category or age 9 or 10
    if "rekrutt" in cat_lower or "gutter 9" in cat_lower or "jenter 9" in cat_lower:
        # Extract prefix (Gutter/Jenter)
        if cat_lower.startswith("gutter"):
            return "Gutter 10"
        elif cat_lower.startswith("jenter"):
            return "Jenter 10"

    # Check for numeric age and collapse if <=10
    # Pattern: "Gutter 9", "Jenter 10", etc.
    parts = category.split()
    if len(parts) >= 2:
        prefix = parts[0]  # "Gutter" or "Jenter"
        try:
            age_str = parts[1]
            # Handle "6-8" format
            if "-" in age_str:
                # Take the highest age in range
                ages = [int(a) for a in age_str.split("-")]
                age = max(ages)
            else:
                age = int(age_str)

            if age <= 10:
                return f"{prefix} 10"
        except (ValueError, IndexError):
            pass  # Not a numeric age, return as-is

    # Return unchanged if no collapsing needed
    return category


def extract_events_from_isonen_csv(csv_path: Path) -> list[EventRow]:
    """Extract unique category+event combinations from Isonen CSV.

    Args:
        csv_path: Path to Isonen format CSV file

    Returns:
        List of unique EventRow objects with collapsed categories
    """
    events: set[tuple[str, str]] = set()

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for row in reader:
            try:
                # Extract category and event
                category = row["Klasse"].strip()
                event = row["Øvelse"].strip()

                # Collapse category
                collapsed_category = collapse_category(category)

                # Add to set (automatically deduplicates)
                events.add((collapsed_category, event))

            except (KeyError, ValueError):
                # Skip invalid rows
                continue

    # Convert to sorted list of EventRow objects
    event_rows = [EventRow(cat, evt) for cat, evt in sorted(events)]
    return event_rows


def export_to_opentrack_xlsx(
    events: list[EventRow],
    output_path: Path,
    meet_name: str = "Track Meet"
) -> None:
    """Export events to OpenTrack XLSX format.

    Creates an Excel file with the following structure:
    - Sheet 1: Events list with Category and Event columns
    - Formatted headers and auto-sized columns

    Args:
        events: List of EventRow objects
        output_path: Path where XLSX file will be written
        meet_name: Name of the meet (used in sheet title)
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    # Write header row with formatting
    headers = ["Category", "Event"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write event data
    for row_idx, event in enumerate(events, start=2):
        ws.cell(row=row_idx, column=1, value=event.category)
        ws.cell(row=row_idx, column=2, value=event.event)

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        # Check all cells in column
        for row in ws[column_letter]:
            try:
                cell_value = str(row.value)
                max_length = max(max_length, len(cell_value))
            except:
                pass

        # Set column width (add padding)
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_path)


def export_isonen_csv_to_opentrack_xlsx(
    csv_path: Path,
    output_path: Path,
    meet_name: str = "Track Meet"
) -> int:
    """Convert Isonen CSV to OpenTrack XLSX format (full pipeline).

    Args:
        csv_path: Path to Isonen format CSV file
        output_path: Path where XLSX file will be written
        meet_name: Name of the meet (used in sheet title)

    Returns:
        Number of unique events exported
    """
    # Extract events
    events = extract_events_from_isonen_csv(csv_path)

    # Export to XLSX
    export_to_opentrack_xlsx(events, output_path, meet_name)

    return len(events)
